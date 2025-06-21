import os
import uuid
import json
import tempfile
import pytesseract
from PIL import Image

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.shortcuts import render
from django.views.decorators.http import require_GET

from langchain_community.vectorstores import FAISS
from langchain_huggingface import HuggingFaceEmbeddings
from langchain.text_splitter import CharacterTextSplitter
from langchain_mistralai import ChatMistralAI
from langchain.chains.combine_documents import create_stuff_documents_chain
from langchain_core.prompts import PromptTemplate
from langchain.docstore.document import Document
from langchain.chains import LLMChain

import logging
logger = logging.getLogger(__name__)
# ========== Config ==========
from dotenv import load_dotenv
load_dotenv()  
ms_api_key = os.getenv('MistralAI_API_TOKEN')

embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/paraphrase-MiniLM-L3-v2")
llm = ChatMistralAI(model="mistral-small", api_key=ms_api_key)

GENERAL_PROMPT = PromptTemplate.from_template("""
Use the following context (which may include tabular data) to answer the question. The context comes from documents the user has uploaded.
Try to interpret tables and values correctly.
When analyzing spreadsheet data:
1. Identify all [number] tagged values
2. For comparison questions (>50, <30 etc):
   - First extract all number-value pairs
   - Then apply the comparison mathematically
   - Finally return names matching the condition
 
                                              
Context:
{context}

Question: {question}

Answer in a clear and concise manner. If the question asks about specific documents, make sure to only use information from those documents.
If you can't find the answer in the provided context, say "I couldn't find that information in your documents".
""")

document_selector_prompt = PromptTemplate.from_template("""You are an intelligent assistant. Given a user's question and a list of available document names,
decide what action to take and which documents to use.

Return a JSON object in one of these exact formats:

1. To list which documents to use to answer the question:
{{"type": "ask_question", "documents": ["doc1.pdf", "doc2.pdf"]}}

2. To check which documents mention a concept:
{{"type": "file_lookup", "query": "concept_to_search"}}
                                                        


Special Handling for Spreadsheets:
- If the question contains numerical terms (marks, scores, values, max, min, average, sum), prioritize Excel files
- For questions about specific sheets or tabs, include the sheet name in brackets: "data.xlsx[Sheet1]"
- For cell references (A1, B2, etc.), include the Excel file that likely contains them
                                                        
For Excel-related questions:
- If asking about cell locations (e.g., "where is 101") → use ask_question
- If asking numerical analysis (max, min, average) → use ask_question
- If searching for existence of values → use ask_question
                                                        
Important Rules:
- Only return valid JSON
- Use double quotes for all strings
- Do not include any additional text or explanations
- The response must be parseable by json.loads()
                                                        
Spreadsheet Question Indicators:
- Contains terms like: marks, scores, values, data, numbers, calculate, sum, average, max, min
- OR includes words like: present, visited, done, marked, checked, attended, complete, submitted, exists, shown, selected
- OR is asking if a specific word or name is in the sheet (e.g., "Is Mumbai marked", "Was Ali present", "Did Aavez attend")
- OR is comparing values, even if not numeric (e.g., "Who is present", "What is the status of Delhi")
- References cells (A1, B2), columns (Column A), or rows (Row 3)
- Asks about tabular data relationships

User Question:
{question}

Available Documents:
{document_names}""")

ALLOWED_EXTENSIONS = {'pdf', 'docx', 'txt', 'pptx', 'jpg', 'jpeg', 'png', 'bmp', 'xlsx', 'md', 'ppt','py', 'csv', 'json', 'html', 'xml','css','java','js','ts','c','cpp','go','php','ruby','swift','kotlin','sql'}

# ========== Views ==========

def index(request):
    return render(request, 'index.html')

@csrf_exempt
@require_GET
def get_uploaded_files(request):
    try:
        documents_data = request.session.get('documents', {})
        files = [
            {
                'id': file_id,
                'name': data['name'],
                'type': data['type']
            }
            for file_id, data in documents_data.items()
        ]
        print(f"Retrieved files: {files}")
        return JsonResponse({'status': 'success', 'files': files})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt
@require_POST
def upload_file(request):
    try:
        if not request.FILES:
            logger.warning("No files uploaded")
            return JsonResponse({'status': 'error', 'message': 'No files uploaded'}, status=400)

        if 'documents' not in request.session:
            request.session['documents'] = {}

        documents_data = request.session['documents']
        results = []
        logger.info("Processing uploaded files",documents_data)
        
        for file in request.FILES.getlist('files'):
            file_name = file.name
            file_ext = os.path.splitext(file_name)[1][1:].lower()
            if file_ext not in ALLOWED_EXTENSIONS:
                return JsonResponse({'status': 'error', 'message': f'Unsupported file type: {file_ext}'}, status=400)
            logger.info(f"Processing file: {file_name} with extension: {file_ext}")
            
            existing_file_id = None
            for file_id, file_data in documents_data.items():
                if file_data['name'].lower() == file_name.lower():
                    existing_file_id = file_id
                    break

            file_id = existing_file_id if existing_file_id else str(uuid.uuid4())

            if existing_file_id:
                try:
                    faiss_files = [
                        f"{documents_data[existing_file_id]['path']}.faiss",
                        f"{documents_data[existing_file_id]['path']}.pkl"
                    ]
                    for faiss_file in faiss_files:
                        if os.path.exists(faiss_file):
                            os.remove(faiss_file)
                except Exception as e:
                    logger.error(f"Error deleting existing FAISS files: {e}")

            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_ext}") as tmp_file:
                for chunk in file.chunks():
                    tmp_file.write(chunk)
                tmp_file_path = tmp_file.name

            if os.path.getsize(tmp_file_path) == 0:
                os.unlink(tmp_file_path)
                results.append({
                    'id': file_id,
                    'name': file_name,
                    'type': file_ext,
                    'status': 'error',
                    'message': 'File is empty'
                })
                continue

            text = extract_text_from_file(tmp_file_path, file_ext)
            os.unlink(tmp_file_path)
            print("text",text)
            if text:
                text_splitter = CharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
                print(f"split text from {file_name}: {text_splitter}...")
                chunks = text_splitter.split_text(text)
                print(f"Number of chunks created from {chunks}: {len(chunks)}")
                documents = [Document(page_content=chunk, metadata={"source": file_name}) for chunk in chunks]
                vectorstore = FAISS.from_documents(documents, embeddings)

                FAISS_SAVE_DIR = os.path.join('media', 'vectorstores')
                os.makedirs(FAISS_SAVE_DIR, exist_ok=True)
                faiss_path = os.path.join(FAISS_SAVE_DIR, file_id)
                vectorstore.save_local(faiss_path)

                documents_data[file_id] = {
                    'name': file_name,
                    'type': file_ext,
                    'path': faiss_path
                }

                results.append({
                    'id': file_id,
                    'name': file_name,
                    'type': file_ext,
                    'status': 'success',
                    'action': 'replaced' if existing_file_id else 'uploaded'
                })
            else:
                results.append({
                    'id': file_id,
                    'name': file_name,
                    'type': file_ext,
                    'status': 'error',
                    'message': 'Could not extract text'
                })

        request.session.modified = True
        logger.info("File upload and processing completed successfully")
        return JsonResponse({'status': 'success', 'files': results})

    except Exception as e:
        logger.error(f"Error processing file: {str(e)}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': 'An error occurred while processing your request.'}, status=500)

@csrf_exempt
@require_POST
def ask_question(request):
    try:
        data = json.loads(request.body)
        question = data.get('question')

        if not question:
            return JsonResponse({'status': 'error', 'message': 'Question required'}, status=400)

        documents_data = request.session.get('documents', {})
        if not documents_data:
            return JsonResponse({'status': 'error', 'message': 'No uploaded documents found.'}, status=400)

        doc_names = [doc['name'] for doc in documents_data.values()]
        
        # Get the LLM response with the fixed prompt
        selector_response = (document_selector_prompt | llm).invoke({
            "question": question,
            "document_names": "\n".join(doc_names)
        })
        
        # Clean and parse the response
        response_text = selector_response.content.strip()
        
        # Handle markdown code blocks if present
        if response_text.startswith('```json'):
            response_text = response_text[7:-3].strip()
        elif response_text.startswith('```'):
            response_text = response_text[3:-3].strip()
        
        try:
            json_start = response_text.index('{')
            json_end = response_text.rindex('}') + 1
            response_text = response_text[json_start:json_end]
            selection = json.loads(response_text)
            selection = json.loads(response_text)
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse LLM response: {response_text}")
            return JsonResponse({
                'status': 'error', 
                'message': 'The assistant returned an invalid response format.',
                'llm_response': response_text
            }, status=400)

        if selection.get("type") == "file_lookup":
            keyword = selection.get("query", "")
            matching_content = []
            
            for doc in documents_data.values():
                vectorstore = FAISS.load_local(doc['path'], embeddings, allow_dangerous_deserialization=True)
                results = vectorstore.similarity_search(keyword, k=3)  # Get top 2 relevant chunks
                filtered_content = [result.page_content for result in results if keyword.lower() in result.page_content.lower()]
                if filtered_content:
                    # Store both the document name and relevant content
                    matching_content.append({
                        'document': doc['name'],
                        'content': [result.page_content for result in results]
                    })
            for item in matching_content:
                print(f"Found in {item['document']}: {item['content']}")
            if not matching_content:
                return JsonResponse({'status': 'success', 'answer': f'No information found about "{keyword}" in your documents.'})
            
            # Format the response with both document names and content
            doc_list = "\n".join(f"- {item['document']}" for item in matching_content)
            response = {
                'status': 'success',
                'answer': f'Information about "{keyword}" was found in:\n{doc_list}',
            }
            print(f"Response for file lookup: {response}")
            return JsonResponse(response)

        elif selection.get("type") == "ask_question":
            selected_docs = selection.get("documents", [])
            print(f"Selected documents for question: {selected_docs}")
            target_docs = [doc for doc in documents_data.values() if doc['name'] in selected_docs]
            
            if not target_docs:
                return JsonResponse({'status': 'success', 'answer': "I couldn't find the requested documents."})

            retrieved_docs = []
            for doc in target_docs:
                vectorstore = FAISS.load_local(doc['path'], embeddings, allow_dangerous_deserialization=True)
                retrieved_docs.extend(vectorstore.similarity_search(question, k=4))

            chain = create_stuff_documents_chain(llm=llm, prompt=GENERAL_PROMPT)
            answer = chain.invoke({
                "context": retrieved_docs,
                "question": question
            })

            return JsonResponse({'status': 'success', 'answer': answer})

        else:
            return JsonResponse({'status': 'error', 'message': 'Unexpected response type from assistant.'}, status=400)

    except Exception as e:
        logger.error(f"Error in ask_question: {str(e)}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
@csrf_exempt
@require_POST
def reset_session(request):
    try:
        if 'documents' in request.session:
            del request.session['documents']
            request.session.modified = True
        return JsonResponse({'status': 'success', 'message': 'Session reset'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# ========== Helper Function ==========

def extract_text_from_file(file_path, file_type):
    try:
        if file_type == 'pdf':
            from langchain_community.document_loaders import PyPDFLoader
            loader = PyPDFLoader(file_path)
            docs = loader.load()
            return "\n\n".join(doc.page_content for doc in docs)

        elif file_type == 'docx':
            from langchain_community.document_loaders import Docx2txtLoader
            loader = Docx2txtLoader(file_path)
            docs = loader.load()
            return docs[0].page_content if docs else None

        elif file_type in ['pptx', 'ppt']:
            from langchain_community.document_loaders import UnstructuredPowerPointLoader
            loader = UnstructuredPowerPointLoader(file_path)
            docs = loader.load()
            return docs[0].page_content if docs else None

        elif file_type in ['jpg', 'jpeg', 'png', 'bmp', 'gif']:
            img = Image.open(file_path)
            return pytesseract.image_to_string(img)
        elif file_type == 'xlsx':
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            wb = load_workbook(file_path, data_only=True)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"\n\nSheet: {sheet}\n"
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        if cell is not None:
                            col_letter = get_column_letter(col_idx)
                            text += f"{col_letter}{row_idx}: {cell}\n"
            return text

        else:  # .txt, .md etc.
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
    except Exception as e:
        print(f"Error processing file [{file_path}]: {e}")
        return None

@csrf_exempt
@require_POST
def delete_file(request):
    try:
        data = json.loads(request.body)
        file_id = data.get('file_id')

        if not file_id:
            return JsonResponse({'status': 'error', 'message': 'File ID required'}, status=400)

        documents_data = request.session.get('documents', {})
        if file_id not in documents_data:
            return JsonResponse({'status': 'error', 'message': 'File not found'}, status=404)

        file_data = documents_data[file_id]
        try:
            faiss_files = [
                f"{file_data['path']}.faiss",
                f"{file_data['path']}.pkl"
            ]
            for faiss_file in faiss_files:
                if os.path.exists(faiss_file):
                    os.remove(faiss_file)
        except Exception as e:
            print(f"Error deleting FAISS files: {e}")

        del documents_data[file_id]
        request.session['documents'] = documents_data
        request.session.modified = True

        return JsonResponse({'status': 'success', 'message': 'File deleted'})

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)